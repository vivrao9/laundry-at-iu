#! python3

import csv
import datetime
import openpyxl
import requests
import slack

#spoof user-agent else IU cyberinfrastructure blocks our request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36"
}

#establish date/time
now = datetime.datetime.now()
date = now.strftime("%m-%d-%Y")
time = now.strftime("%H:%M:%S")

slack_token = 'xoxp-16039322196-299587575255-732518687043-8e4fb1223689101e81ebc3e556d82449'
client = slack.WebClient(token=slack_token)

wb = openpyxl.load_workbook('C:\\Users\\Vivek Rao\\Desktop\\IDS\\Enterprise Desk\\Laundry Status Scraping\\Laundry XHR Syntax.xlsx')
sheet = wb['Nomenclature']

try:
        for i in range(2, 94):
                center = sheet.cell(row=i, column=2).value
                bldg = sheet.cell(row=i, column=4).value
                floor = sheet.cell(row=i, column=6).value
            
                res = requests.get(sheet.cell(row=i, column=7).value, headers=headers).json()

                #print each key/value to a .csv
                for j in range(len(res)):
                        dictionary = res[j]
                        dataList = list(dictionary.values())

                        #define variables
                        machineID = dataList[0]
                        machineType = dataList[1]
                        machineNumber = dataList[2]
                        machineStatus = dataList[3]
                        machineAvailability = dataList[4]
                        minutesRemaining = dataList[5]
                        notification = dataList[6]
                        isRequested = dataList[7]
                        isLoggedIn = dataList[8]

                        #write to a .csv
                        outputFile = open('C:\\Users\\Vivek Rao\\Desktop\\IDS\\Enterprise Desk\\Laundry Status Scraping\\Laundry Statuses.csv', 'a', newline='\n')
                        outputWriter = csv.writer(outputFile)
                        outputWriter.writerow([date, time, center, bldg, floor, machineID, machineType, machineNumber, machineStatus, machineAvailability, minutesRemaining, notification, isRequested, isLoggedIn])
                outputFile.close()
                        
        client.chat_postMessage(
                  channel="CMJFFFYAX",
                  text="Laundry Larry has successfully extracted information from every washer and dryer on campus. :tada:",
                  icon_emoji=":robot_face:",
                  username="Laundry Larry"
                                )       

except:
                client.chat_postMessage(
                  channel="CMJFFFYAX",
                  text="Save the Queen! Something\'s gone wrong! @vivrao",
                  link_names=1,
                  icon_emoji=":robot_face:",
                  username="Laundry Larry"
                                )
